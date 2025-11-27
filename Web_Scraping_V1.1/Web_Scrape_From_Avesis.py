import requests
from bs4 import BeautifulSoup
from openpyxl import workbook, load_workbook
import pandas as pd
import numpy as np

url = ["https://avesis.ktu.edu.tr/zaferyavuz/yayinlar","https://avesis.ktu.edu.tr/murat_aykut/yayinlar","https://avesis.ktu.edu.tr/baharhatipoglu/yayinlar",
       "https://avesis.ktu.edu.tr/selenguven/yayinlar","https://avesis.ktu.edu.tr/osivaz/yayinlar","https://avesis.ktu.edu.tr/vasif/yayinlar","https://avesis.ktu.edu.tr/bekir/yayinlar",
       "https://avesis.ktu.edu.tr/seymademir/yayinlar","https://avesis.ktu.edu.tr/ckose/yayinlar","https://avesis.ktu.edu.tr/bustubioglu/yayinlar","https://avesis.ktu.edu.tr/sedatgormus/yayinlar",
       "https://avesis.ktu.edu.tr/muhammedkilic/yayinlar","https://avesis.ktu.edu.tr/cakiro/yayinlar","https://avesis.ktu.edu.tr/ulutas/yayinlar","https://avesis.ktu.edu.tr/dincer/yayinlar",
       "https://avesis.ktu.edu.tr/gultahaoglu/yayinlar","https://avesis.ktu.edu.tr/ekinci/yayinlar","https://avesis.ktu.edu.tr/pehlivan/yayinlar","https://avesis.ktu.edu.tr/cevhers/yayinlar",
       "https://avesis.ktu.edu.tr/burakaydin/yayinlar","https://avesis.ktu.edu.tr/gulutas/yayinlar"]
headers={
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 OPR/123.0.0.0",
    "Accept-Language":"tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer":"https://www.google.com/",
    "Connection":"keep-alive"
}

wb = load_workbook('Avesis_Makaleler.xlsx')
ws = wb.active
ws.delete_rows(1, ws.max_row)

eklenen_kayitlar = set()

for link in url:
    response = requests.get(link, headers= headers)
    soup = BeautifulSoup(response.text, "html.parser")

    table = soup.find_all("div", class_="item-body")[0]
    makales = table.find_all("div", class_="pub-item with-icon")
    for makale in makales:
        titles = [t.text.strip() for t in makale.find_all("strong")]
        yazarlar = [y.text.strip() for y in makale.find_all("a", class_="authorsRichText")]

        satir = tuple(titles + yazarlar)
        
        if satir in eklenen_kayitlar:
            continue
        eklenen_kayitlar.add(satir)
        ws.append(list(satir))
    wb.save("Avesis_Makaleler.xlsx")